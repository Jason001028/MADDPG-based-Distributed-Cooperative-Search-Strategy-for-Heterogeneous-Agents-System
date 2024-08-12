import numpy as np
import matplotlib.pyplot as plt
import pygame
import random
import gym
import math
import copy
import sys 
from easydict import EasyDict as edict
import csv
import os
import openpyxl as op
import pandas as pd

env_params = edict({
    'grid_size': 20,
    'n_agents':  3,
    'observation_dim': 35,
    'action_dim': 5,
    'clip_obs': False,
    'max_timesteps': 100,
    })
#easydict模块，简化字典调用


# def initialize_cover_csv(path):
#     file_name = os.path.join(path, 'training_cover_data.csv')
#     keys = ['step' , 'agent0', 'agent1', 'agent2']
#     with open(file_name, 'a', newline='') as f:
#         writer = csv.writer(f)
#         writer.writerow(keys)
#     return file_name

class Gridworld(gym.Env):
    def __init__(self, agent_num = 3, obstacles = None):
        # Initialize pygame
        pygame.init()
        self.agent_num = agent_num
        self.seed = 10
        self.save_fig_time = 0
        # Set up window and font
        self.font = pygame.font.SysFont(None, 30)
        self.window_size = (400, 400)
        self.window = pygame.display.set_mode(self.window_size)
        pygame.display.set_caption('Gridworld')
        self.save_max_num = 10
        #安全距离
        self.safe_dis = 5
        self.total_clear = 0
        self.smog_count = 0
        #初始迷雾量
        self.smog_initial_count = 0

        # Initialize gridworld
        #视野范围2
        self.viewrange = 2
        #最早是20，我们缩小至16
        self.grid_size = 16
        self.num_actions = 5
        #行动空间，5自由度
        self.action_mapping = {0: [0, 0], 1: [-1, 0], 2: [1, 0], 3: [0, -1], 4: [0, 1]}
        self.num_states = self.grid_size * self.grid_size
        
        self.origin_obstacle_states = []
        self.origin_stable_obstacle_states = copy.deepcopy(obstacles)
        self.goal_state = []

        ##agent0,1,2: explorer,postman,surveyor
        self.agent_task_rate = [0.3,0,0.7]

        self.agent_task_viewrange = [1,0,4]

        self.agent_cover_count = [0] * agent_num
        self.agent0_cover = []
        self.agent1_cover = []
        self.agent2_cover = []

        self.agent0_move_count = 0
        self.agent1_move_count = 0
        self.agent2_move_count = 0

        # init obstacle pos
        # for _ in range(60):
        #     x = random.randint(1,self.grid_size-1 )
        #     y = random.randint(1,self.grid_size-1 )
        #     if [x,y] not in self.origin_obstacle_states and [x,y] not in self.goal_state:
        #         self.origin_stable_obstacle_states.append([x,y])
        # init agent pos
        origin_pos_tmp = agent_num
        self.origin_current_state = []
        while origin_pos_tmp > 0:
            x = random.randint(1, self.grid_size - 1)
            y = random.randint(1, self.grid_size - 1)
            if [x, y] not in self.origin_obstacle_states and [x, y] not in self.goal_state and [x,y] not in self.origin_current_state and [x,y] not in self.origin_stable_obstacle_states:
                self.origin_current_state.append([x,y])
                origin_pos_tmp -= 1
        # init goal pos\
        self.agent0_cover.append(self.origin_current_state[0])
        self.agent1_cover.append(self.origin_current_state[1])
        self.agent2_cover.append(self.origin_current_state[2])
        self.obstacle_movement_prob = 0.05
        self.max_step = 200

        # define for RL
        cp_obs_space = 3 * self.grid_size * self.grid_size + 2 * len(self.goal_state) + 4 + 1 + 2 # map 3 * 20 * 20 + 15
        # self.observation_space = gym.spaces.Box(low = -1, high=1, shape = (4, self.grid_size, self.grid_size,))
        self.observation_space = gym.spaces.Box(low = -1, high=1, shape = (cp_obs_space,))
        self.action_space = [
gym.spaces.Box(low = -1, high=1, shape = (1,)) for _ in range(self.agent_num) # 0,1 for move, 2,3 for 20 load\unload, 4,5 for 40 load\unload
        ]
        self.reset()

##地图迷雾函数
    def smog(self):
        self.goal_state = []
        self.smog_count = 0
        goal_smog_tmp = (self.grid_size-1)^2
        while goal_smog_tmp > 0:
            for x in range(0, self.grid_size):
                for y in range(0, self.grid_size):
                    if [x, y] not in self.origin_obstacle_states and [x, y] not in self.goal_state and [x, y] not in self.origin_current_state and [x,y] not in self.origin_stable_obstacle_states:
                        self.goal_state.append([x, y])
                        self.smog_count += 1
                        goal_smog_tmp -= 1
        return self.smog_count

    # def sample_goals(self):
    #     self.goal_state = []
    #     goal_pos_tmp = self.agent_num
    #     while goal_pos_tmp > 0:
    #         x = random.randint(1, self.grid_size - 1)
    #         y = random.randint(1, self.grid_size - 1)
    #         if [x, y] not in self.origin_obstacle_states and [x, y] not in self.goal_state and
    #         [x,y] not in self.origin_current_state and [x,y] not in self.origin_stable_obstacle_states:
    #             self.goal_state.append([x,y])
    #             goal_pos_tmp -= 1
    
    def get_distance(self, pos1, pos2):
        return abs(pos1[0] - pos2[0]) + abs(pos1[1] - pos2[1])#返回绝对距离

    def reset(self):
        self.trajectory = [[] for _ in range(self.agent_num)]
        self.stable_obstacle_states = copy.deepcopy(self.origin_stable_obstacle_states)
        self.current_state = copy.deepcopy(self.origin_current_state)
        self.obstacle_states = copy.deepcopy(self.origin_obstacle_states)
        # self.sample_goals()
        #不生成目标，而是迷雾
        # #smog_count：剩余迷雾单元格数量
        self.smog_count = self.smog()
        self.smog_realtime_count = self.smog()
        self.smog_initial_count = self.smog_realtime_count
        self.agent_cover_count = [0] * self.agent_num

        self.cur_step = 0
        #0代表未达到，1代表达到 先暂时全注释掉
        # self.get_goal = [0] * self.agent_num
        self.last_dis = [0] * self.agent_num
        self.expect_smog_dis = [2] * self.agent_num
        for i in range(self.agent_num):
            for j in range(self.agent_num):
                self.last_dis[i] = min(self.last_dis[i], self.get_distance(self.current_state[i], self.goal_state[j]))
        state = [self.get_state(i) for i in range(self.agent_num)]
        self.clear_smog(i)
        plot_path = 'saved_models'
        # csv_file_name = initialize_cover_csv(plot_path)
        return state

###迷雾消除机制
#功能：从列表中清除指定迷雾，并返回一个清除数
    def clear_smog(self, i):
        my_x, my_y = self.current_state[i]
        # data = np.array(self.goal_state)
        # np.delete(data,[my_x, my_y])
        # self.goal_state=list(data)
        # print(self.goal_state)
        count_remove = 0
        if i == 0:
            for y in range(my_y - self.agent_task_viewrange[i], my_y + self.agent_task_viewrange[i] + 1):
                for x in range(my_x - self.agent_task_viewrange[i], my_x + self.agent_task_viewrange[i] + 1):
                    if [x, y] in self.goal_state:
                        self.goal_state.remove([x, y])
                        # if [x, y] not in self.agent1_cover and [x, y] not in self.agent2_cover:
                        if [x, y] not in self.agent2_cover:
                            self.agent0_cover.append([x, y])
                            self.agent_cover_count[i] += 1
                        count_remove = count_remove+1
                    else:
                        pass
        if i == 1:
            if [my_x , my_y] in self.goal_state:
                self.goal_state.remove([my_x, my_y])
                if [my_x , my_y] not in self.agent0_cover and [my_x, my_y] not in self.agent2_cover:
                    self.agent1_cover.append([my_x , my_y])
                    self.agent_cover_count[i] += 1
                count_remove = count_remove + 1
        if i == 2:
            for y in range(my_y - self.agent_task_viewrange[i], my_y + self.agent_task_viewrange[i] + 1):
                for x in range(my_x - self.agent_task_viewrange[i], my_x + self.agent_task_viewrange[i] + 1):
                    if [x, y] in self.goal_state:
                        # for  and [x,y] not in self.origin_stable_obstacle_states
                        self.goal_state.remove([x, y])
                        if [x, y] not in self.agent0_cover and [x, y] not in self.agent1_cover:
                            self.agent2_cover.append([x, y])
                            self.agent_cover_count[i] += 1
                        count_remove = count_remove+1
                    else:
                        pass

        return count_remove

        # del self.goal_state [my_x,my_y]




    def get_state(self, i):###获取智能体当前位置，及其观测
        total_obs = [] 
        total_obs.append(self.cur_step / self.max_step)
        # agent pos
        my_x, my_y = self.current_state[i]
        total_obs.append(my_x/self.grid_size)
        total_obs.append(my_y/self.grid_size)
        for j in range(self.agent_num):
            x, y = self.current_state[j]
            total_obs.append(x/self.grid_size)
            total_obs.append(y/self.grid_size)
            if len(self.goal_state) >= 3:
                goal_x, goal_y = self.goal_state[j]
            else:
                goal_x, goal_y = [8, 8]
            total_obs.append(goal_x/self.grid_size)
            total_obs.append(goal_y/self.grid_size)
            total_obs.append((my_x - goal_x) / self.grid_size)
            total_obs.append((my_y - goal_y) / self.grid_size)
            total_obs.append((my_x - x) / self.grid_size)
            total_obs.append((my_y - y) / self.grid_size)
        
        # get available action
        total_obs.extend(self.get_availabel_action(i))
        agent_id = [0, 0, 0]
        agent_id[i] = 1
        total_obs.extend(agent_id)
        # is in goal
        is_in_goal = 1 if self.current_state[i] in self.goal_state else 0
        total_obs.append(is_in_goal)
        return total_obs
    
    def get_availabel_action(self, agent_id):
        direction = list(self.action_mapping.values())
        available_action = [1] * 5
        for i, direc in enumerate(direction):
            new_x, new_y = self.current_state[agent_id][0] + direc[0], self.current_state[agent_id][1] + direc[1]
            if [new_x, new_y] in self.stable_obstacle_states or ([new_x, new_y] in self.stable_obstacle_states) or new_x < 0 or new_y < 0 or new_x >= self.grid_size or new_y >= self.grid_size:
                available_action[i] = 0
            for j in range(self.agent_num):
                if new_x == self.current_state[j][0] and new_y == self.current_state[j][1]:
                    available_action[i] = 0
        available_action[0] = 1
        return available_action
    
    def savefig(self, name):
        self.save_fig_time  += 1
        if self.save_max_num > self.save_fig_time:
            pygame.image.save(self.window, f"path_Saving_{name}_{self.save_fig_time}.png")

    def parse_action(self, actions):

        new_actions = np.argmax(actions, axis= -1)

        return new_actions

##在actor.py被调用   单回合智能体行动步骤
    def step(self, t, actions):
        actions = self.parse_action(actions)
        self.cur_step += 1
        for action in actions:
            if action < 0 or action >= self.num_actions:#出现异常
                raise Exception('Invalid action: {}'.format(action))
        assert len(actions) == self.agent_num, f'actions length is {len(actions)}, agent_num {self.agent_num}'
        # 会根据智能体数目产生负奖励
        rewards = [-0.005] * self.agent_num
        count_oneclear_total = [0, 0, 0]
        cita = [1, 1, 1]
        done = 0
        for i, action in enumerate(actions):
            #更新奖励系数cita
            #达到任务量，会限制探索，但给予大量奖励
            if self.agent_cover_count[i] >= self.smog_initial_count * self.agent_task_rate[i]:
                rewards[i] -= 1
                cita[i] = 0.1
            #清除迷雾，并计数
            count = self.clear_smog(i)
            rewards[i] += cita[i]*count
            # print(cita[i]*count)
            count_oneclear_total[i] += count
            self.total_clear += count
            valid_actions = self.get_availabel_action(i)
            if(self.smog_realtime_count > count):
                self.smog_realtime_count -= count
            ##什么都不做，产生负奖励
            if valid_actions[action] == 0:
                if self.agent_cover_count[i] < self.agent_task_rate[i]*200:
                    rewards[i] -= 15
            else:
                #次态位置坐标
                next_state = [self.current_state[i][0] + self.action_mapping[action][0], self.current_state[i][1] + self.action_mapping[action][1]]
                # shorted_dis = min(self.get_distance(next_state, self.goal_state[j]) for j in range(self.agent_num))
                #机器人间的距离
                shorted_dis = min(self.get_distance(next_state, self.current_state[j]) for j in range(self.agent_num))
                #如果距离过近，会产生一个负奖励
                avoid_rate = 10
                tele_rate = 5
                if i == 0 or i == 2:
                    if shorted_dis <= self.safe_dis:
                        if shorted_dis == 5:
                            rewards[i] -= 0.05*avoid_rate
                        if shorted_dis == 4:
                            rewards[i] -= 0.45*avoid_rate
                        if shorted_dis == 3:
                            rewards[i] -= 0.65*avoid_rate
                        if shorted_dis == 2:
                            rewards[i] -= 0.85*avoid_rate
                        if shorted_dis <= 1:
                            rewards[i] -= 1*avoid_rate
                        else:
                            rewards[i] += 0.15

                if i == 1:
                    if shorted_dis == 5:
                        rewards[i] += 0.05 * tele_rate
                    if shorted_dis == 4:
                        rewards[i] += 0.45 * tele_rate
                    if shorted_dis == 3:
                        rewards[i] += 0.65 * tele_rate
                    if shorted_dis == 2:
                        rewards[i] += 0.85 * tele_rate
                    if shorted_dis <= 1:
                        rewards[i] += 1 * tele_rate
                    else:
                        rewards[i] -= 0.25
                # if shorted_dis < self.last_dis[i]:
                #     rewards[i] -= 0.01
                #     self.last_dis[i] = shorted_dis

                # #与迷雾间的距离，鼓励自发探索行为
                # if i == 0 or i == 2:
                #     shorted_smog_dis = min(self.get_distance(self.current_state[i], self.goal_state[j])
                #     for j in range(len(self.goal_state)))
                #     if shorted_smog_dis <= self.expect_smog_dis[i]:
                #         rewards[i] += 0.15
                #         # print('有效启发探索')
                #     else:
                #         rewards[i] -= 0.15


                self.current_state[i] = next_state
                self.trajectory[i].append(self.current_state[i])
                if self.current_state[i] in self.goal_state:
                    idx = self.goal_state.index(self.current_state[i])
                    # print(idx)
                    # if self.get_goal[idx] == 0:
                    #     rewards[i] += 1
                    #     self.get_goal[idx] = 1
        # done, total_r = self.get_is_done()
        # #全体获得集体奖励与集体惩罚，当探索量到达规定量后，合计奖励才为正值
        # total_r = (sum(count_oneclear_total)-count_oneclear_total[1]) * 0.15 + count_oneclear_total[2]*0.3 + (30 - self.smog_realtime_count) * 0.02

        #全体获得集体奖励与集体惩罚，当探索量到达规定量后，合计奖励才为正值
        total_r = (sum(count_oneclear_total) - count_oneclear_total[1]) * 0.65 - self.smog_realtime_count * 0.1 + (
                    self.smog_initial_count - self.smog_realtime_count) * 0.3
        escape_rate = (self.agent_cover_count[0]/self.agent_cover_count[2])/(self.agent_task_rate[0]/self.agent_task_rate[2])
        if 0.25 < escape_rate < 4:
            total_r += 10*math.atan(1/(escape_rate-1))
            if 0.5 < escape_rate < 2:
                total_r += 10*math.atan(1/(escape_rate-1))
            if escape_rate == 1:
                total_r += 20
        else:
            if escape_rate != 0:
                if escape_rate > 1:
                    total_r -= 5*math.atan(escape_rate-1)
                if escape_rate < 1:
                    total_r -= 5*math.atan(1/escape_rate)
            if escape_rate == 0:
                total_r -= 6


        if count_oneclear_total.count(0) == len(count_oneclear_total) and self.smog_realtime_count > 20:
            total_r -= 5

        rewards = [r+total_r for r in rewards]
        info = [] if total_r != 1 else [True]
        sub_agent_obs = []
        sub_agent_reward = []
        sub_agent_done = []
        sub_agent_info = []

        # #返回cover值,用于最后制图
        # bg = op.load_workbook(r"saved_models/agent_cover.xlsx")
        # sheet = bg["Sheet1"]  # “Sheet1”表示将数据写入到excel文件的sheet1下
        # row_num = 1
        # bgplus = bg.active
        # for i in range(1,t+1):
        #     if bgplus.cell(row=i,column=1).value != None:
        #          row_num += 1
        # for i in range(1, len(self.agent_cover_count)+1):
        #     sheet.cell(row_num, i, self.agent_cover_count[i-1])  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
        # bg.save("saved_models/agent_cover.xlsx")  # 对文件进行保存
        # bg.close()

        for i in range(self.agent_num):
            sub_agent_obs.append(self.get_state(i))
            sub_agent_reward.append(rewards[i])
            sub_agent_done.append(done)
            sub_agent_info.append(info)
#这里返回值容易error
        # print(count_oneclear_total)
        return escape_rate,self.agent_cover_count, sub_agent_obs, sub_agent_reward, sub_agent_done, sub_agent_info


    def get_is_done(self):
            if self.cur_step >= self.max_step:
                return True, 0
            if self.total_clear != self.smog_initial_count:
                return False, 0
            return True, 1

    def reward_func(self):
        rewards = [0] * self.agent_num
        return rewards

    ##重新显示并绘制窗口图
    def render(self, escape_rate, reward, done, save_path_name = None):
        # Clear window
        self.window.fill((200, 200, 200))
        row_size = self.window_size[0] / self.grid_size
        col_size = self.window_size[1] / self.grid_size
        
        # Draw grid lines
        for i in range(self.grid_size+1):
            pygame.draw.line(self.window, (0, 0, 0), (0, i*col_size), (self.window_size[0], i*col_size), 1)
            pygame.draw.line(self.window, (0, 0, 0), (i*row_size, 0), (i*row_size, self.window_size[1]), 1)

        # Draw obstacles 障碍物
        for obstacle_state in self.origin_stable_obstacle_states:
            color = (0, 0, 0)
            # if obstacle_state in self.stable_obstacle_states:
            #     color = (0, 0, 0)
            pygame.draw.rect(self.window, color, (obstacle_state[1]*row_size, obstacle_state[0]*col_size, row_size, col_size))

        # Draw cover 已清除区域
        for grid1 in self.agent0_cover:
            color = (143, 170, 220)
            pygame.draw.rect(self.window, color, (grid1[1] * row_size, grid1[0] * col_size, row_size, col_size))
        for grid2 in self.agent1_cover:
            color = (238, 213, 142)
            pygame.draw.rect(self.window, color, (grid2[1] * row_size, grid2[0] * col_size, row_size, col_size))
        for grid3 in self.agent2_cover:
            color = (142, 215, 238)
            pygame.draw.rect(self.window, color, (grid3[1] * row_size, grid3[0] * col_size, row_size, col_size))
        for grid1 in self.agent0_cover:
            color = (143, 170, 220)
            pygame.draw.rect(self.window, color, (grid1[1] * row_size, grid1[0] * col_size, row_size, col_size))

        # Draw smog 地图迷雾
        for goal in self.goal_state:
            color = (112, 102, 104)
            # if self.get_goal[self.goal_state.index(goal)] == 0:
            #     color = (0, 255, 0)
            pygame.draw.rect(self.window, color, (goal[1]*row_size, goal[0]*col_size, row_size, col_size))

        # # Draw goal state 目标地点
        # for goal in self.goal_state:
        #     color = (255, 0, 0)
        #     # if self.get_goal[self.goal_state.index(goal)] == 0:
        #     #     color = (0, 255, 0)
        #     pygame.draw.rect(self.window, color, (goal[1]*row_size, goal[0]*col_size, row_size, col_size))

        # Draw agent 按照异构智能体分配
        for i in range(self.agent_num):
            x, y = self.current_state[i][0], self.current_state[i][1]
            if i == 0:
                 pygame.draw.rect(self.window, (29, 122, 235), (y*row_size, x*col_size, row_size, col_size)) #蓝：explorer
            if i == 1:
                 pygame.draw.rect(self.window, (210, 98, 23), (y*row_size, x*col_size, row_size, col_size)) #橙：postman
            if i == 2:
                 pygame.draw.rect(self.window, (28, 149, 188), (y*row_size, x*col_size, row_size, col_size)) #青：surveyor

        # # Draw trajectory绘制轨迹  建议注释
        # for agent_traj in self.trajectory:
        #     for point in agent_traj:
        #         # pygame.draw.circle(self.window, (111, 25, 230), ((0.5+point[1])*row_size, (0.5+point[0])*col_size), 5, width=1)
        #         pygame.draw.circle(self.window, (111, 25, 230),((0.5 + point[1])*row_size,(0.5+point[0])*col_size),5,width=1)
        # Draw reward and done status
        for i in range(0,3):
             reward[i]=round(reward[i],2)
        escape_rate = round(escape_rate, 4)
        reward_text = self.font.render('Reward: {}'.format(reward), True, (0, 0, 0))
        done_text = self.font.render('Agent_cover: {}'.format(self.agent_cover_count), True, (0, 0, 0))
        escape_rate_text = self.font.render('escape_rate: {}'.format(escape_rate), True, (0,0,0))

        # sum_reward_text = self.font.render('Done: {}'.format(sum_reward), True, (0, 0, 0))
        self.window.blit(reward_text, (10, self.window_size[1]-40))
        self.window.blit(done_text, (10, self.window_size[1]-70))
        self.window.blit(escape_rate_text, (10, self.window_size[1]-100))

        # self.window.blit(sum_reward_text, (10, self.window_size[1] - 100))

        # Update display
        pygame.display.update()
        if save_path_name is not None:
            pygame.image.save(self.window, save_path_name)

        # Check for quit event
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                sys.exit()

    def euclidean_distance(self, point1, point2):
        return abs(point1[0] - point2[0]) + abs(point1[1] - point2[1])

if __name__ == "__main__":
    import time
    #会反复刷新
    def test(env, num_steps=1000):
        for i in range(10):
            a_num = env.agent_num_callback()
            total_reward = 0
            env.reset()
            env.render(0, False)
            for _ in range(num_steps):
                action = [np.random.randint(0, env.num_actions - 1 ) for i in range(env.agent_num)]
                next_state, reward, done, _ = env.step(action)
                env.render(sum(reward), done[0])
                total_reward += sum(reward)
                time.sleep(0.1)
                if done[0]:
                    break
            print('Total reward: {}'.format(total_reward))
    env = Gridworld()
    test(env)